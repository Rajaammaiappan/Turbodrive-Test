# -*- coding: utf-8 -*-
"""
CAIRO-Assist  -  Inspection Criteria Comparator  (RR / ALTEN internal tool)
===========================================================================
Upload two OR MORE Rolls-Royce AeroManager / Pinpoint MHTML snapshots of the
SAME maintenance task. Every document is compared against every other document
(no baseline - the comparison is symmetric). The tool shows a light-coloured
heat-map plus an all-pairs agreement matrix, and exports an Excel workbook that
matches whatever you have filtered on screen.

Environment (confirmed available - no internet pip needed):
    Python 3.9.12 (Anaconda)   Flask 1.1.2 / Werkzeug 2.0.3
    beautifulsoup4 4.11.1  lxml 4.9.1  openpyxl 3.0.10
Run:
    "C:\\ProgramData\\Anaconda3\\python.exe" cairo_assist.py
    (or double-click the .bat launcher)
"""
import os, re, io, email, json, difflib, threading, webbrowser, time, traceback, uuid
import sqlite3, getpass, socket, csv
from collections import Counter
from datetime import datetime

from flask import Flask, request, jsonify, send_file, Response
from werkzeug.utils import secure_filename

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
PORT = 5001
app.config['MAX_CONTENT_LENGTH'] = 120 * 1024 * 1024   # 120 MB total upload

STORE = {}   # last comparison model, keyed by token

# ============================================================================
#  USAGE LOG DATABASE  --  EDIT THIS ONE LINE to set where the log is saved.
#  Every time the tool is used, a row is written to this SQLite (.db) file.
#  It is created automatically if it does not exist. Examples:
#     LOG_DB_PATH = r"\\portfolioeng_nlr\EFS\CAIRO-Assist\usage_log.db"
#     LOG_DB_PATH = r"C:\Users\u8531675\OneDrive - Rolls-Royce\CAIRO-Assist_usage_log.db"
#  Default = a file named 'cairo_assist_log.db' next to this script.
# ============================================================================
LOG_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cairo_assist_log.db")

def _db_connect():
    d = os.path.dirname(LOG_DB_PATH)
    if d and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)
    con = sqlite3.connect(LOG_DB_PATH, timeout=10)
    con.execute("""CREATE TABLE IF NOT EXISTS usage_log(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT, doc_type TEXT, action TEXT, user TEXT, machine TEXT,
        task_no TEXT, task_title TEXT, documents TEXT, doc_count INTEGER,
        rows INTEGER, note TEXT)""")
    return con

def current_user():
    for key in ('USERNAME', 'USER'):
        v = os.environ.get(key)
        if v:
            return v
    try:
        return getpass.getuser()
    except Exception:
        return 'unknown'

def current_machine():
    return os.environ.get('COMPUTERNAME') or socket.gethostname() or 'unknown'

def log_event(doc_type, action, task_no='', task_title='', documents=None,
              doc_count=0, rows=0, note=''):
    """Write one audit row. Never raises - logging must not break the tool."""
    try:
        docs = ' | '.join(documents or [])
        con = _db_connect()
        con.execute("""INSERT INTO usage_log
            (timestamp, doc_type, action, user, machine, task_no, task_title,
             documents, doc_count, rows, note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), doc_type, action,
             current_user(), current_machine(), task_no or '', task_title or '',
             docs, doc_count, rows, note))
        con.commit(); con.close()
    except Exception as e:
        # last-resort: print to console so the tool keeps working
        print('[usage-log] could not write log: %s' % e)

# ------------------------------------------------------- light colour palette
COLORS = {
    'MATCH':     'C6EFCE',   # light green
    'PARTIAL':   'FFF2CC',   # light yellow
    'THRESHOLD': 'FCE4D6',   # light orange
    'MISSING':   'FFC7CE',   # light red / pink
    'UNIQUE':    'E4DFEC',   # light purple
    'NA':        'F2F2F2',   # light grey (criterion not applicable here)
}
STATUS_ORDER = ['MATCH', 'PARTIAL', 'THRESHOLD', 'MISSING', 'UNIQUE']

# ================================================================= PARSING
def clean(s):
    return re.sub(r'\s+', ' ', s.replace('\xa0', ' ')).strip()

def load_main_html_from_bytes(raw):
    msg = email.message_from_bytes(raw)
    subject = None
    for k, v in msg.items():
        if k.lower() == 'subject':
            subject = v
    best = ""
    for part in msg.walk():
        if part.get_content_type() == 'text/html':
            payload = part.get_payload(decode=True)
            if payload:
                h = payload.decode('utf-8', 'ignore')
                if len(h) > len(best):
                    best = h
    return subject, best

def norm_disp(s):
    dl = s.lower()
    if dl.startswith('accept'):
        return 'Accept'
    if dl.startswith('reject'):
        return 'Reject'
    if 'repair' in dl or 'refer to' in dl or 'remove' in dl or 'blend' in dl:
        return 'Repair'
    return s[:40]

NAV_H4 = {'print preview', 'print options', 'bookmark manager', 'delete ?',
    'close dialog without saving?', 'export control acknowledgement', 'unlock publications',
    'challenge code', 'enter the response code:', 'content not found', 'list of annotations',
    'select csn link to open', 'select document to open', 'select sns link to open',
    'installation context links', 'graphic hotspot list', 'referenced by', 'amm-references',
    'ipc-csn list', 'figures', 'graphic list', 'release log', 'table of contents',
    'list of figures', 'reference', 'references'}

def parse_bytes(raw, fallback_name):
    subject, html = load_main_html_from_bytes(raw)
    soup = BeautifulSoup(html, 'lxml')
    vp = soup.select_one('.viewPage') or soup

    # ---- task title / number ----
    task_title = None
    hf = soup.select_one('.pgHeaderFooter')                      # Trent 1000 exam style
    if hf:
        for c in hf.find_all(['td', 'th']):
            t = clean(c.get_text(' '))
            if t and 'Manual' not in t and 'Export' not in t and len(t) < 60:
                task_title = t
                break
    if not task_title:                                           # else first meaningful heading
        for hh in vp.find_all(['h1', 'h2', 'h3', 'h4']):
            t = clean(hh.get_text(' '))
            if t and t.lower() not in NAV_H4 and not t.upper().startswith('SUBTASK') \
               and t.lower() not in ('description', 'procedure', 'examinations, tests, and checks') \
               and len(t) < 110 and (' - ' in t or 'Examine' in t or 'Repair' in t or len(t) > 12):
                task_title = t
                break
    task_no = None
    m = re.search(r'\b(\d{2}-\d{2}-\d{2}-\d{3}-\d{3})\b', vp.get_text(' '))
    if m:
        task_no = m.group(1)
    if not task_no:
        m2 = re.search(r'\((\d{2}-\d{2}-\d{2}),', vp.get_text(' '))
        if m2:
            task_no = m2.group(1)

    # ---- walk content: keep ALL tables, with their real section/type ----
    records = []
    cur_h4 = None       # descriptive <h4> section  ("Braze the Part", "Consumable Materials", "Examine the ...")
    cur_colon = None    # short line ending ':'  (XWB area, or exam condition)
    cur_sub = None      # "SUBTASK 72-41-12-xxx"
    for el in vp.descendants:
        name = getattr(el, 'name', None)
        if name is None:
            continue
        if name in ('h1', 'h2', 'h3', 'h4') and el.find('table') is None:
            t = clean(el.get_text(' '))
            if not t:
                continue
            if t.upper().startswith('SUBTASK'):
                cur_sub = t
            elif t.lower() not in NAV_H4 and len(t) < 90:
                cur_h4 = t
        elif name in ('div', 'span', 'p') and el.find('table') is None:
            t = clean(el.get_text(' '))
            if t.endswith(':') and len(t) < 80 and not t.lower().startswith('revdate') \
               and not t.lower().startswith(('model', 'effectivity')):
                cur_colon = t[:-1].strip()
        if name == 'table' and 'src-table' in (el.get('class') or []):
            rows = el.find_all('tr')
            if not rows:
                continue
            head = [clean(c.get_text(' ')).upper() for c in rows[0].find_all(['td', 'th'])]
            hjoin = ' | '.join(head)
            is_dla = (len(head) >= 3 and head[0].startswith('DAMAGE')
                      and 'LIMIT' in hjoin and 'ACTION' in hjoin)
            is_proc = ('PROCEDURE' in (head[0] if head else '') and 'RELATED DATA' in hjoin)

            if is_dla:
                # XWB criteria: DAMAGE | LIMIT | ACTION
                area = cur_colon or cur_h4 or '(root)'
                last_dmg = ''
                for tr in rows[1:]:
                    cells = [clean(c.get_text(' ')) for c in tr.find_all(['td', 'th'])]
                    if len(cells) < 3:
                        continue
                    dmg, limit, action = cells[0], cells[1], cells[-1]
                    if dmg:
                        last_dmg = dmg
                    if limit and action:
                        records.append({'subtask': area, 'condition': dmg or last_dmg,
                                        'criterion': limit, 'disposition': norm_disp(action),
                                        'kind': 'criteria'})
            elif is_proc:
                # Repair procedure steps: PROCEDURE | RELATED DATA
                area = cur_h4 or cur_colon or '(root)'
                for tr in rows[1:]:
                    cells = [clean(c.get_text(' ')) for c in tr.find_all(['td', 'th'])]
                    if not cells or not cells[0]:
                        continue
                    proc = cells[0]
                    related = cells[-1] if len(cells) > 1 else ''
                    records.append({'subtask': area, 'condition': cur_sub or '',
                                    'criterion': proc, 'disposition': (related[:60] or 'Step'),
                                    'kind': 'procedure'})
            else:
                # everything else: examination criteria OR materials / tools / parts lists.
                # Decide per row from the last cell: a real disposition -> criteria, else a list value.
                area = cur_h4 or cur_colon or '(root)'
                for tr in rows:
                    cells = [clean(c.get_text(' ')) for c in tr.find_all(['td', 'th'])]
                    if len(cells) < 2 or not cells[0]:
                        continue
                    key = cells[0] if len(cells) == 2 else ' '.join(cells[:-1])
                    val = cells[-1]
                    vl = val.lower()
                    if vl.startswith(('accept', 'reject')) or 'repair' in vl or 'refer to' in vl:
                        records.append({'subtask': area, 'condition': cur_colon or '',
                                        'criterion': key, 'disposition': norm_disp(val),
                                        'kind': 'criteria'})
                    else:
                        if key.upper() in ('DAMAGE', 'LIMIT', 'ACTION', 'PROCEDURE', 'ENGINE'):
                            continue
                        records.append({'subtask': area, 'condition': '',
                                        'criterion': key, 'disposition': val[:60],
                                        'kind': 'reference'})
    return {'subject': subject or fallback_name, 'task_no': task_no,
            'task_title': task_title, 'records': records}

# ================================================================= MATCHING
def norm_area(s):
    s = s.lower()
    s = re.sub(r'\bhp turbine\b', '', s)
    s = re.sub(r'\bblades\b', 'blade', s)
    s = re.sub(r'\brotor blade\b', 'blade', s)
    s = re.sub(r'\binitially\b', '', s)
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def _dec(s):
    # unify decimal separators: "1,00" (comma) == "1.00" (period)
    return re.sub(r'(?<=\d)\s*,\s*(?=\d)', '.', s)

def norm_crit(s):
    s = _dec(s.lower())
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'refer to.*', '', s)
    s = re.sub(r'[^a-z0-9. ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip(' .')

def numbers(s):
    return re.findall(r'\d+\.?\d*', _dec(s))

def related(a, b):
    """Do these two criteria describe the SAME check (ignoring threshold/wording)?"""
    na, nb = norm_crit(a['criterion']), norm_crit(b['criterion'])
    if na == nb:
        return True
    return difflib.SequenceMatcher(None, na, nb).ratio() > 0.6

def crit_key(r):
    return (norm_crit(r['criterion']), r['disposition'], tuple(numbers(r['criterion'])))

# ================================================================= ALL-PAIRS MODEL
AREA_STOP = {'examine', 'the', 'and', 'for', 'of', 'a', 'an', 'initially', 'do', 'test',
             'part', 'area', 'to', 'in', 'on', 'with'}

def _awords(name):
    ws = re.findall(r'[a-z]+|\d+', name.lower())
    words = set(w for w in ws if w.isalpha() and w not in AREA_STOP)
    nums = set(w for w in ws if w.isdigit())
    return words, nums

def suggest_matches(parsed, existing=None):
    """Suggest cross-file inspection areas that look like the SAME item under a
    different item number / wording (max matching words). User confirms before merge."""
    existing = existing or []
    already = set()
    for grp in existing:
        for n in grp:
            already.add(n)
    per = []
    for p in parsed:
        seen = []
        for r in p['records']:
            if r['subtask'] not in seen:
                seen.append(r['subtask'])
        per.append(seen)

    # best partner per (source area, other file)
    edges = []
    for fi in range(len(per)):
        for a in per[fi]:
            if a in already:
                continue
            wa, na = _awords(a)
            if not wa:
                continue
            for fj in range(len(per)):
                if fj == fi:
                    continue
                if a in per[fj]:
                    continue                       # identical name already aligns
                best = None
                for b in per[fj]:
                    if b in already:
                        continue
                    wb, nb = _awords(b)
                    if not wb:
                        continue
                    uni = wa | wb
                    ws = len(wa & wb) / len(uni) if uni else 0
                    if ws < 0.5:
                        continue
                    key = (round(ws, 3), len(na & nb),
                           round(difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio(), 3))
                    if best is None or key > best[0]:
                        best = (key, b)
                if best:
                    edges.append((best[0], (fi, a), (fj, best[1])))

    # union-find to group equivalent areas across >=2 files
    parent = {}
    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(x, y):
        parent[find(x)] = find(y)
    escore = {}
    for (key, na_, nb_) in edges:
        union(na_, nb_)
        g = None
        escore[(na_, nb_)] = key[0]
    groups = {}
    for (key, na_, nb_) in edges:
        r = find(na_)
        groups.setdefault(r, {'members': set(), 'score': 0})
        groups[r]['members'].add(na_); groups[r]['members'].add(nb_)
        groups[r]['score'] = max(groups[r]['score'], key[0])

    out = []
    for g in groups.values():
        members = sorted(g['members'], key=lambda m: (m[0], m[1]))
        files = set(m[0] for m in members)
        if len(members) < 2 or len(files) < 2:
            continue
        names = [m[1] for m in members]
        wsets = [ _awords(n)[0] for n in names ]
        shared = set.intersection(*wsets) if wsets else set()
        out.append({'members': [{'file': m[0], 'name': m[1]} for m in members],
                    'names': names, 'shared': sorted(shared), 'score': round(g['score'], 2)})
    out.sort(key=lambda s: -s['score'])
    return out[:40]

def apply_aliases(parsed, aliases):
    """Relabel record areas so confirmed-equivalent items share one canonical name."""
    canon = {}
    for grp in aliases or []:
        names = [n for n in grp if n]
        if len(names) < 2:
            continue
        c = min(names, key=len)                    # canonical = shortest name in the group
        for n in names:
            canon[n] = c
    if not canon:
        return parsed, set()
    newp = []
    for p in parsed:
        recs = [dict(r, subtask=canon.get(r['subtask'], r['subtask'])) for r in p['records']]
        newp.append(dict(p, records=recs))
    return newp, set(canon.values())


def build_model(parsed, aliases=None):
    parsed, merged_areas = apply_aliases(parsed, aliases)
    names = [(p['subject'] or 'file')[:45] for p in parsed]
    ndoc = len(parsed)

    # ---- GLOBAL clustering: align criteria across ALL areas/documents by text.
    # Area names differ between engine types (Trent 1000 vs XWB), so criteria are
    # matched by wording. If two criteria sit in a user-confirmed same area, boost the
    # match so "Item 1 ..." and "Item 2 ..." merge into a single row.
    clusters = []   # each: {rep:rec, members:{di:rec}}
    for di in range(ndoc):
        for r in parsed[di]['records']:
            best, bcl = 0.0, None
            for cl in clusters:
                if di in cl['members']:
                    continue
                na, nb = norm_crit(cl['rep']['criterion']), norm_crit(r['criterion'])
                if not na or not nb:
                    continue
                sc = 1.0 if na == nb else difflib.SequenceMatcher(None, na, nb).ratio()
                if r['subtask'] == cl['rep']['subtask'] and r['subtask'] in merged_areas:
                    sc = min(1.0, sc + 0.2)
                if sc > 0.6 and sc > best:
                    best, bcl = sc, cl
            if bcl is not None:
                bcl['members'][di] = r
            else:
                clusters.append({'rep': r, 'members': {di: r}})

    rows = []
    co = [[0] * ndoc for _ in range(ndoc)]
    ag = [[0] * ndoc for _ in range(ndoc)]
    for cl in clusters:
        present = cl['members']
        keys = {di: crit_key(r) for di, r in present.items()}
        consensus = Counter(keys.values()).most_common(1)[0][0]
        cons_txt, cons_disp, cons_nums = consensus
        only_one = len(present) == 1

        cells = []
        for di in range(ndoc):
            if di in present:
                r = present[di]; k = keys[di]
                if only_one:
                    st = 'UNIQUE'
                elif k == consensus:
                    st = 'MATCH'
                elif k[0] == cons_txt and k[1] == cons_disp and k[2] != cons_nums:
                    st = 'THRESHOLD'
                else:
                    st = 'PARTIAL'
                cells.append({'text': r['criterion'] + '  \u2192 ' + r['disposition'], 'status': st})
            else:
                cells.append({'text': '\u2014', 'status': 'NA' if only_one else 'MISSING'})

        pres_idx = list(present.keys())
        for x in range(len(pres_idx)):
            for y in range(x + 1, len(pres_idx)):
                i, j = pres_idx[x], pres_idx[y]
                co[i][j] += 1; co[j][i] += 1
                if keys[i] == keys[j]:
                    ag[i][j] += 1; ag[j][i] += 1

        first = present[pres_idx[0]]
        pretty = Counter([r['subtask'] for r in present.values()]).most_common(1)[0][0]
        cond = Counter([r['condition'] for r in present.values()]).most_common(1)[0][0]
        rep_txt = next((r['criterion'] for di, r in present.items() if keys[di] == consensus),
                       first['criterion'])
        rows.append({'area': pretty, 'condition': cond, 'criterion': rep_txt, 'cells': cells,
                     'merged': pretty in merged_areas})

    # keep the heat-map tidy: group rows by inspection area for display
    order = {}
    for r in rows:
        order.setdefault(r['area'], len(order))
    rows.sort(key=lambda r: order[r['area']])

    summary = []
    for di in range(ndoc):
        c = {s: 0 for s in STATUS_ORDER}; c['present'] = 0
        for r in rows:
            st = r['cells'][di]['status']
            if st in c:
                c[st] += 1
            if st in ('MATCH', 'PARTIAL', 'THRESHOLD', 'UNIQUE'):
                c['present'] += 1
        summary.append(c)

    pairs = []
    for i in range(ndoc):
        rowp = []
        for j in range(ndoc):
            rowp.append(None if i == j else (round(100.0 * ag[i][j] / co[i][j]) if co[i][j] else None))
        pairs.append(rowp)

    return {'names': names, 'task_no': parsed[0]['task_no'], 'task_title': parsed[0]['task_title'],
            'rows': rows, 'summary': summary, 'pairs': pairs, 'co': co,
            'generated': datetime.now().strftime('%d %b %Y  %H:%M')}

# ================================================================= FILTER (shared by UI + download)
def filter_rows(rows, statuses, query):
    sset = set(statuses) if statuses else set(STATUS_ORDER)
    q = (query or '').lower().strip()
    out = []
    for r in rows:
        cstat = [c['status'] for c in r['cells']]
        if not any(s in sset for s in cstat):
            continue
        if q:
            hay = (r['area'] + ' ' + r['condition'] + ' ' + r['criterion'] + ' ' +
                   ' '.join(c['text'] for c in r['cells'])).lower()
            if q not in hay:
                continue
        out.append(r)
    return out

# ================================================================= EXCEL
def write_excel(model, rows):
    wb = Workbook()
    thin = Side(style='thin', color='C8CED8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hdr_fill = PatternFill('solid', fgColor='305496')          # header still readable, text white
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)

    ws = wb.active
    ws.title = 'Heatmap'
    titles = ['Inspection Area', 'Condition', 'Criterion'] + model['names']
    for j, t in enumerate(titles, 1):
        c = ws.cell(1, j, t); c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    for j, w in enumerate([26, 16, 46] + [40] * len(model['names']), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    r, last = 2, None
    if not rows:
        ws.cell(2, 1, 'No rows matched the current filter.').font = Font(name='Arial', italic=True, size=10)
    for row in rows:
        ws.cell(r, 1, row['area'] if row['area'] != last else '').alignment = wrap
        last = row['area']
        ws.cell(r, 2, row['condition']).alignment = wrap
        ws.cell(r, 3, row['criterion']).alignment = wrap
        for k, cell in enumerate(row['cells']):
            xc = ws.cell(r, 4 + k, cell['text'])
            xc.fill = PatternFill('solid', fgColor=COLORS[cell['status']])
            xc.alignment = wrap; xc.font = Font(name='Arial', size=9); xc.border = border
        for j in range(1, 4):
            ws.cell(r, j).font = Font(name='Arial', size=9); ws.cell(r, j).border = border
        r += 1
    ws.freeze_panes = 'D2'
    if r > 2:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(titles)), r - 1)

    # Summary
    ss = wb.create_sheet('Summary')
    ss.cell(1, 1, 'Task').font = Font(name='Arial', bold=True)
    ss.cell(1, 2, "%s  (%s)" % (model['task_title'] or '', model['task_no'] or '')).font = Font(name='Arial')
    ss.cell(2, 1, 'Rows exported').font = Font(name='Arial', bold=True)
    ss.cell(2, 2, len(rows)).font = Font(name='Arial')
    heads = ['Document', 'Match', 'Partial', 'Threshold diff', 'Missing', 'Unique', 'Criteria present']
    for j, h in enumerate(heads, 1):
        c = ss.cell(4, j, h); c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    for i, (nm, cnt) in enumerate(zip(model['names'], model['summary'])):
        rr = 5 + i
        ss.cell(rr, 1, nm).font = Font(name='Arial', size=10)
        for j, key in enumerate(['MATCH', 'PARTIAL', 'THRESHOLD', 'MISSING', 'UNIQUE', 'present'], 2):
            cc = ss.cell(rr, j, cnt[key]); cc.alignment = center; cc.border = border; cc.font = Font(name='Arial', size=10)
    for j, w in enumerate([40, 9, 9, 14, 9, 9, 16], 1):
        ss.column_dimensions[get_column_letter(j)].width = w
    # legend
    lr = 7 + len(model['names'])
    ss.cell(lr, 1, 'Legend').font = Font(name='Arial', bold=True)
    for i, (st, lab) in enumerate([('MATCH', 'Match - all documents agree'),
                                   ('PARTIAL', 'Partial - wording differs, same intent'),
                                   ('THRESHOLD', 'Threshold diff - numeric limit differs'),
                                   ('MISSING', 'Missing - criterion absent in this document'),
                                   ('UNIQUE', 'Unique - criterion only in this document'),
                                   ('NA', 'Not applicable')]):
        cc = ss.cell(lr + 1 + i, 1, lab)
        cc.fill = PatternFill('solid', fgColor=COLORS[st]); cc.font = Font(name='Arial', size=10)

    # Pairwise agreement matrix
    pm = wb.create_sheet('Pairwise')
    pm.cell(1, 1, 'All-pairs agreement (%) - share of shared criteria that match exactly').font = Font(name='Arial', bold=True)
    for j, nm in enumerate(model['names'], 2):
        c = pm.cell(3, j, nm); c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    for i, nm in enumerate(model['names']):
        rr = 4 + i
        c = pm.cell(rr, 1, nm); c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(wrap_text=True); c.border = border
        for j in range(len(model['names'])):
            v = model['pairs'][i][j]
            cc = pm.cell(rr, 2 + j, '\u2014' if v is None else v)
            cc.alignment = center; cc.border = border; cc.font = Font(name='Arial', size=10)
            if v is not None:
                shade = 'C6EFCE' if v >= 90 else ('FFF2CC' if v >= 70 else 'FFC7CE')
                cc.fill = PatternFill('solid', fgColor=shade)
    pm.column_dimensions['A'].width = 34
    for j in range(len(model['names'])):
        pm.column_dimensions[get_column_letter(2 + j)].width = 18

    # Details (full text, unfiltered)
    ds = wb.create_sheet('Details (all)')
    for j, t in enumerate(titles, 1):
        c = ds.cell(1, j, t); c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    for j, w in enumerate([26, 16, 46] + [40] * len(model['names']), 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    for i, row in enumerate(model['rows'], 2):
        ds.cell(i, 1, row['area']).alignment = wrap
        ds.cell(i, 2, row['condition']).alignment = wrap
        ds.cell(i, 3, row['criterion']).alignment = wrap
        for k, cell in enumerate(row['cells']):
            ds.cell(i, 4 + k, cell['text']).alignment = wrap
    ds.freeze_panes = 'D2'

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio

# ================================================================= ROUTES
@app.route('/')
def index():
    return Response(INDEX_HTML, mimetype='text/html')

@app.route('/compare', methods=['POST'])
def compare():
    try:
        files = [f for f in request.files.getlist('files') if f and f.filename]
        if len(files) < 2:
            return jsonify({'ok': False, 'error': 'Please choose at least 2 MHTML files.'}), 400
        orig_names = [f.filename for f in files]
        parsed = [parse_bytes(f.read(), secure_filename(f.filename)) for f in files]
        empties = [p['subject'] for p in parsed if not p['records']]
        if len([p for p in parsed if p['records']]) < 2:
            return jsonify({'ok': False, 'error': 'Fewer than 2 files contained inspection criteria. '
                            'Make sure each snapshot has the task open in Pinpoint before saving as MHTML.'}), 400
        model = build_model(parsed)
        token = uuid.uuid4().hex[:12]
        STORE[token] = model
        STORE[token]['orig_names'] = orig_names
        STORE[token]['parsed'] = parsed          # keep for re-compute with confirmed matches
        STORE[token]['aliases'] = []
        log_event('Manuals', 'compare', task_no=model['task_no'], task_title=model['task_title'],
                  documents=orig_names, doc_count=len(orig_names), rows=len(model['rows']),
                  note='%d criteria' % len(model['rows']))
        payload = {k: model[k] for k in ('names', 'task_no', 'task_title', 'rows', 'summary', 'pairs', 'generated')}
        payload['token'] = token
        payload['suggestions'] = suggest_matches(parsed)
        payload['warnings'] = ['"%s" had no criteria (empty or a different task).' % e for e in empties]
        return jsonify({'ok': True, 'model': payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/recompute', methods=['POST'])
def recompute():
    """Re-run the comparison after the user confirms item/area matches (aliases)."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        entry = STORE.get(data.get('token'))
        if not entry or 'parsed' not in entry:
            return jsonify({'ok': False, 'error': 'Result expired - run the comparison again.'}), 404
        aliases = data.get('aliases') or []
        parsed = entry['parsed']
        model = build_model(parsed, aliases)
        token = data['token']
        model['parsed'] = parsed
        model['orig_names'] = entry.get('orig_names')
        model['aliases'] = aliases
        STORE[token] = model
        log_event('Manuals', 'recompute', task_no=model['task_no'], task_title=model['task_title'],
                  documents=entry.get('orig_names'), doc_count=len(model['names']),
                  rows=len(model['rows']), note='merged %d item group(s)' % len(aliases))
        payload = {k: model[k] for k in ('names', 'task_no', 'task_title', 'rows', 'summary', 'pairs', 'generated')}
        payload['token'] = token
        payload['suggestions'] = suggest_matches(parsed, aliases)
        payload['merged_count'] = len(aliases)
        return jsonify({'ok': True, 'model': payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/download', methods=['POST'])
def download():
    data = request.get_json(force=True, silent=True) or {}
    model = STORE.get(data.get('token'))
    if not model:
        return jsonify({'ok': False, 'error': 'Result expired - run the comparison again.'}), 404
    rows = filter_rows(model['rows'], data.get('statuses'), data.get('query'))
    bio = write_excel(model, rows)
    log_event('Manuals', 'download', task_no=model.get('task_no'), task_title=model.get('task_title'),
              documents=model.get('orig_names') or model.get('names'), doc_count=len(model['names']),
              rows=len(rows), note='filtered export (%d of %d rows)' % (len(rows), len(model['rows'])))
    fname = 'CAIRO-Assist_%s_%s.xlsx' % (model['task_no'] or 'task', datetime.now().strftime('%Y%m%d_%H%M%S'))
    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    try:
        return send_file(bio, as_attachment=True, attachment_filename=fname, mimetype=mime)
    except TypeError:
        return send_file(bio, as_attachment=True, download_name=fname, mimetype=mime)

@app.route('/logs')
def logs():
    try:
        con = _db_connect()
        cur = con.execute("""SELECT id,timestamp,doc_type,action,user,machine,task_no,
                             documents,doc_count,rows,note FROM usage_log ORDER BY id DESC LIMIT 500""")
        data = cur.fetchall(); con.close()
    except Exception as e:
        return Response('<h3>Could not read log DB</h3><pre>%s</pre>' % e, mimetype='text/html')
    cols = ['id', 'timestamp', 'doc type', 'action', 'user', 'machine', 'task', 'documents', 'docs', 'rows', 'note']
    trs = ''
    for r in data:
        trs += '<tr>' + ''.join('<td>%s</td>' % ('' if v is None else str(v)) for v in r) + '</tr>'
    html = LOGS_HTML.replace('__PATH__', LOG_DB_PATH).replace('__COUNT__', str(len(data)))
    html = html.replace('__HEAD__', ''.join('<th>%s</th>' % c for c in cols)).replace('__ROWS__', trs)
    return Response(html, mimetype='text/html')

@app.route('/logs.csv')
def logs_csv():
    con = _db_connect()
    cur = con.execute("""SELECT id,timestamp,doc_type,action,user,machine,task_no,task_title,
                         documents,doc_count,rows,note FROM usage_log ORDER BY id DESC""")
    rows = cur.fetchall(); con.close()
    sio = io.StringIO(); w = csv.writer(sio)
    w.writerow(['id', 'timestamp', 'doc_type', 'action', 'user', 'machine', 'task_no',
                'task_title', 'documents', 'doc_count', 'rows', 'note'])
    w.writerows(rows)
    bio = io.BytesIO(sio.getvalue().encode('utf-8-sig')); bio.seek(0)
    fname = 'CAIRO-Assist_usage_%s.csv' % datetime.now().strftime('%Y%m%d_%H%M%S')
    try:
        return send_file(bio, as_attachment=True, attachment_filename=fname, mimetype='text/csv')
    except TypeError:
        return send_file(bio, as_attachment=True, download_name=fname, mimetype='text/csv')

# ================================================================= FRONT-END
INDEX_HTML = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>CAIRO-Assist</title>
<style>
 :root{--bg:#eef2f9;--surface:#fff;--surface2:#f5f8fd;--border:#c8d4e8;--accent:#1a4fad;--accent2:#1d6fdb;
   --text:#1e293b;--muted:#5b6675;--mono:'IBM Plex Mono',Consolas,monospace;
   --c-match:#C6EFCE;--c-partial:#FFF2CC;--c-thr:#FCE4D6;--c-miss:#FFC7CE;--c-uni:#E4DFEC;--c-na:#F2F2F2;}
 *{box-sizing:border-box}
 body{margin:0;background:var(--bg);color:var(--text);font-family:system-ui,"Segoe UI",sans-serif;font-size:14px}
 .mono{font-family:var(--mono)}
 /* header like TPCR: logo left corner, title, logo right corner, light grey bar */
 header{display:flex;align-items:center;gap:16px;background:linear-gradient(180deg,#f6f8fb,#e9eef6);
   border-bottom:2px solid #cfd8e6;padding:10px 18px}
 .brandbox{width:52px;height:40px;display:flex;align-items:center;justify-content:center;background:#fff;
   border:1px solid var(--border);border-radius:6px;overflow:hidden;flex:0 0 auto}
 .brandbox img{max-width:100%;max-height:100%;object-fit:contain}
 .header-title{flex:1}
 .header-title h1{margin:0;font-size:20px;color:#12203a}
 .header-title .subtitle{margin:2px 0 0;color:var(--muted);font-size:12px}
 .fallback-alten{font-weight:800;color:#111;font-size:13px}.fallback-alten span{color:#e2001a}
 .fallback-rr{font-weight:800;color:#00205b;font-size:14px;letter-spacing:1px}
 .wrap{max-width:1600px;margin:0 auto;padding:18px 20px}
 .card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;margin-bottom:16px}
 .card h2{margin:0 0 4px;font-size:15px}.card p.hint{margin:0 0 12px;color:var(--muted);font-size:12px}
 .pills{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap}
 .pill{background:var(--surface2);border:1px solid var(--border);border-radius:16px;padding:3px 12px;font-size:12px;color:var(--muted)}
 .pill b{color:var(--accent)}
 .doctypes{display:flex;flex-wrap:wrap;gap:10px 22px;align-items:center;margin-top:6px}
 .dt{display:inline-flex;align-items:center;gap:7px;font-size:14px;cursor:pointer;color:var(--text)}
 .dt input{width:16px;height:16px;accent-color:var(--accent);cursor:pointer}
 .dt.soon{color:#9aa6b6;cursor:not-allowed}
 .dt.soon s{color:#9aa6b6}
 .dt .soon-tag{color:#c05252;font-size:11px;font-style:italic}
 .dt.soon input{cursor:not-allowed}
 .drop{border:2px dashed var(--border);border-radius:10px;background:var(--surface2);padding:22px;text-align:center;cursor:pointer}
 .drop.hi{border-color:var(--accent2);background:#e8f0fe}.drop input{display:none}
 .filelist{margin:12px 0 0;padding:0;list-style:none}
 .filelist li{display:flex;align-items:center;gap:10px;padding:7px 10px;border:1px solid var(--border);border-radius:8px;margin-bottom:6px;background:#fff}
 .filelist li .nm{flex:1;font-size:13px}
 .filelist li .rm{color:#dc2626;cursor:pointer;border:none;background:none;font-size:16px}
 .row{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-top:12px}
 button.primary{background:var(--accent);color:#fff;border:none;border-radius:8px;padding:10px 20px;font-size:14px;cursor:pointer}
 button.primary:disabled{background:#9db4d8;cursor:not-allowed}
 button.ghost{background:#fff;border:1px solid var(--border);border-radius:8px;padding:9px 16px;cursor:pointer}
 #progressWrap{margin-top:12px;display:none}
 #progressBar{height:8px;background:var(--accent2);border-radius:5px;width:0;transition:width .3s}
 #progressMsg{font-size:12px;color:var(--muted);margin-top:5px;display:block}
 .err{background:#fee2e2;border:1px solid #fecaca;color:#991b1b;border-radius:8px;padding:10px 12px;font-size:13px;margin-top:12px}
 .warn{background:#fef3c7;border:1px solid #fde68a;color:#92400e;border-radius:8px;padding:8px 12px;font-size:12px;margin-top:10px}
 #results{display:none}
 .engines{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px}
 .ecard{border:1px solid var(--border);border-radius:8px;padding:10px 12px;background:#fff;min-width:210px;flex:1}
 .ecard h3{margin:0 0 8px;font-size:13px}
 .bar{display:flex;height:12px;border-radius:6px;overflow:hidden;border:1px solid #e2e5e2}.bar span{display:block}
 .counts{display:flex;flex-wrap:wrap;gap:6px 12px;margin-top:8px;font-size:12px;color:#42505f}.counts b{font-variant-numeric:tabular-nums}
 .matrix{border-collapse:collapse;font-size:12px;margin-top:4px}
 .matrix th,.matrix td{border:1px solid var(--border);padding:6px 9px;text-align:center}
 .matrix th{background:var(--surface2);font-weight:600}
 .matrix td.self{background:#eef2f9;color:#9aa6b6}
 .toolbar{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:12px}
 .toolbar input[type=search]{padding:8px 11px;border:1px solid var(--border);border-radius:7px;min-width:250px;font-size:13px}
 .chip{border:1px solid var(--border);background:#fff;border-radius:20px;padding:5px 11px;font-size:12px;cursor:pointer;user-select:none;display:inline-flex;align-items:center;gap:7px}
 .chip .sw{width:12px;height:12px;border-radius:3px;border:1px solid #0002}.chip.off{opacity:.35}
 .dlnote{font-size:11px;color:var(--muted)}
 .grid{border:1px solid var(--border);border-radius:8px;overflow:auto;background:#fff;max-height:70vh}
 table.heat{border-collapse:separate;border-spacing:0;width:100%;font-size:12.5px}
 table.heat th,table.heat td{padding:8px 10px;border-bottom:1px solid #eceef2;border-right:1px solid #eceef2;vertical-align:top;text-align:left}
 table.heat thead th{position:sticky;top:0;background:#eef2f9;color:#22314f;z-index:3;font-weight:600;font-size:12px;border-bottom:1px solid #cfd8e6}
 table.heat tbody th.area{position:sticky;left:0;background:#fbfcfe;z-index:2;font-weight:600;min-width:150px;max-width:180px}
 td.crit{min-width:250px;max-width:330px;color:#26313f}
 td.cell{min-width:200px;max-width:300px;position:relative}
 td.cell .disp{display:inline-block;font-size:10px;font-weight:700;padding:1px 6px;border-radius:4px;background:#0000000f;margin-left:4px}
 .st-MATCH{background:var(--c-match)}.st-PARTIAL{background:var(--c-partial)}.st-THRESHOLD{background:var(--c-thr)}
 .st-MISSING{background:var(--c-miss)}.st-UNIQUE{background:var(--c-uni)}.st-NA{background:var(--c-na);color:#9aa6b6}
 tr.arowtop th.area{border-top:2px solid #d3dcea}tr.arowtop td{border-top:2px solid #eef0f4}
 .flag{position:absolute;top:4px;right:5px;font-size:9px;font-weight:700;padding:1px 5px;border-radius:8px;background:#0000000f;color:#4b5563}
 .msug{border:1px solid var(--border);border-radius:8px;padding:8px 10px;margin-bottom:6px;background:#fff;display:flex;gap:10px;align-items:flex-start}
 .msug input{margin-top:3px;width:16px;height:16px;accent-color:var(--accent)}
 .msug .names{flex:1;font-size:12.5px}
 .msug .nm{display:block;margin:1px 0}
 .msug .fb{display:inline-block;font-size:10px;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:0 6px;margin-right:6px;color:var(--muted)}
 .msug .sc{font-size:11px;color:#2a7d2a;font-weight:700;white-space:nowrap}
 .msug .wd{color:#d11313;font-weight:700}
 .mergedtag{display:inline-block;font-size:9px;font-weight:700;color:var(--accent);border:1px solid #9cc0e2;border-radius:8px;padding:0 5px;margin-left:6px;vertical-align:middle}
 .d{color:#d11313;font-weight:700;text-decoration:underline;text-decoration-color:#d11313}
 .disp.dred{color:#d11313;font-weight:800}
 footer{color:var(--muted);font-size:11px;text-align:center;padding:16px}
</style></head>
<body>
<header>
  <div class="brandbox" title="ALTEN">
    <img src="https://www.alten.com/wp-content/uploads/2019/01/favicon-alten.png" alt="ALTEN"
         onerror="this.outerHTML='<div class=&quot;fallback-alten&quot;>AL<span>T</span>EN</div>'">
  </div>
  <div class="header-title">
    <h1>CAIRO-Assist</h1>
    <p class="subtitle">Criteria Analysis &amp; Inspection Reconciliation &mdash; all-document comparison &middot; Internal Use Only</p>
  </div>
  <div class="brandbox" title="Rolls-Royce">
    <img src="https://www.rolls-royce.com/~/media/Images/R/Rolls-Royce/logo/rebrand-svg-logo.svg" alt="Rolls-Royce"
         onerror="this.outerHTML='<div class=&quot;fallback-rr&quot;>RR</div>'">
  </div>
</header>

<div class="wrap">
  <div class="card" id="docTypeCard">
    <h2 style="color:var(--accent)">HeatMap</h2>
    <p class="hint">Tick <b>Manuals</b> to compare maintenance-manual inspection criteria (Accept/Reject) across
       MHTML snapshots. The other document types are coming soon and will use the same heat-map engine.</p>
    <div class="doctypes">
      <label class="dt"><input type="checkbox" id="dt-manuals" checked onchange="checkReady()"> <b>Manuals</b></label>
      <label class="dt soon"><input type="checkbox" disabled> <s>TV</s> <span class="soon-tag">(coming soon)</span></label>
      <label class="dt soon"><input type="checkbox" disabled> <s>Concession</s> <span class="soon-tag">(coming soon)</span></label>
      <label class="dt soon"><input type="checkbox" disabled> <s>RST &amp; TRM</s> <span class="soon-tag">(coming soon)</span></label>
    </div>
  </div>

  <div class="card" id="inputCard">
    <div class="pills">
      <span class="pill">Step <b>1</b> &middot; Add MHTML files</span>
      <span class="pill">Step <b>2</b> &middot; Run all-pairs comparison</span>
      <span class="pill">Step <b>3</b> &middot; Filter &amp; download</span>
    </div>
    <h2>Select MHTML snapshots (2 or more)</h2>
    <p class="hint">Every document is compared against <b>every other</b> document &mdash; there is no baseline.
       Each file must have the same maintenance task open in Pinpoint.</p>
    <label class="drop" id="drop">
      <input type="file" id="fileInput" accept=".mhtml,.mht" multiple>
      <div><b>Click to choose files</b> or drag &amp; drop here</div>
      <div style="font-size:12px;color:var(--muted);margin-top:4px">.mhtml / .mht &middot; up to 120&nbsp;MB total</div>
    </label>
    <ul class="filelist" id="filelist"></ul>
    <div class="row">
      <button class="primary" id="runBtn" onclick="runTool()" disabled>&#9654; Run comparison</button>
      <button class="ghost" onclick="clearAll()">Clear</button>
      <span class="dlnote" id="dtNote"></span>
    </div>
    <div id="progressWrap"><div id="progressBar"></div><span id="progressMsg">Starting&hellip;</span></div>
    <div id="errBox"></div>
  </div>

  <div class="card" id="results">
    <div class="row" style="justify-content:space-between;margin-top:0">
      <div><h2 id="rTitle" style="margin-bottom:2px">Results</h2>
        <div class="mono" id="rTask" style="color:var(--muted);font-size:12px"></div></div>
      <div style="text-align:right">
        <button class="primary" id="dlBtn">&#8681; Download Excel (filtered)</button>
        <div class="dlnote" id="dlNote"></div>
      </div>
    </div>
    <div id="warnBox"></div>

    <div class="engines" id="engines" style="margin-top:12px"></div>

    <details style="margin-bottom:14px" open>
      <summary style="cursor:pointer;font-weight:600;font-size:13px">All-pairs agreement matrix</summary>
      <div id="matrixWrap" style="overflow:auto;margin-top:8px"></div>
      <div style="font-size:11px;color:var(--muted);margin-top:4px">Each cell = % of criteria shared by both documents that match exactly.</div>
    </details>

    <details id="matchDetails" style="margin-bottom:14px">
      <summary style="cursor:pointer;font-weight:600;font-size:13px">Suggested item matches <span id="matchCount" style="color:var(--muted);font-weight:400"></span></summary>
      <div style="font-size:12px;color:var(--muted);margin:6px 0">These areas look like the same item under a different number/wording. Tick the ones that are the same, then re-compare to merge them into single rows.</div>
      <div id="matchList"></div>
      <div class="row" style="margin-top:8px">
        <button class="primary" id="applyMatchBtn" onclick="applyMatches()">&#10003; Apply matches &amp; re-compare</button>
        <span class="dlnote" id="matchNote"></span>
      </div>
    </details>

    <div class="toolbar">
      <input type="search" id="q" placeholder="Search area, criterion or value (e.g. dimension 18)">
      <span id="chips"></span>
    </div>
    <div class="grid"><table class="heat"><thead><tr id="hrow"></tr></thead><tbody id="tbody"></tbody></table></div>
  </div>
</div>

<footer>&copy; 2026 Alten-Rolls-Royce. All rights reserved. Confidential &ndash; Internal Use Only.
  &nbsp;&middot;&nbsp; <a href="/logs" target="_blank" style="color:var(--accent);text-decoration:none">Usage log</a></footer>

<script>
var chosen=[], MODEL=null;
var STLABEL={MATCH:'Match',PARTIAL:'Partial',THRESHOLD:'Threshold diff',MISSING:'Missing',UNIQUE:'Unique'};
var STCOLOR={MATCH:'var(--c-match)',PARTIAL:'var(--c-partial)',THRESHOLD:'var(--c-thr)',MISSING:'var(--c-miss)',UNIQUE:'var(--c-uni)',NA:'var(--c-na)'};
var ORDER=['MATCH','PARTIAL','THRESHOLD','MISSING','UNIQUE'];
var active={MATCH:1,PARTIAL:1,THRESHOLD:1,MISSING:1,UNIQUE:1};

var drop=document.getElementById('drop'), fi=document.getElementById('fileInput');
fi.addEventListener('change',function(e){addFiles(e.target.files);});
['dragenter','dragover'].forEach(function(ev){drop.addEventListener(ev,function(e){e.preventDefault();drop.classList.add('hi');});});
['dragleave','drop'].forEach(function(ev){drop.addEventListener(ev,function(e){e.preventDefault();drop.classList.remove('hi');});});
drop.addEventListener('drop',function(e){addFiles(e.dataTransfer.files);});

function addFiles(list){
  for(var i=0;i<list.length;i++){var f=list[i];
    if(!/\.(mhtml|mht)$/i.test(f.name))continue;
    if(!chosen.some(function(c){return c.name===f.name&&c.size===f.size;}))chosen.push(f);}
  renderFiles();
}
function removeFile(i){chosen.splice(i,1);renderFiles();}
function clearAll(){chosen=[];fi.value='';renderFiles();document.getElementById('results').style.display='none';document.getElementById('errBox').innerHTML='';}
function manualsOn(){return document.getElementById('dt-manuals').checked;}
function checkReady(){
  var on=manualsOn();
  document.getElementById('inputCard').style.opacity = on ? '1' : '.5';
  document.getElementById('fileInput').disabled = !on;
  var note=document.getElementById('dtNote');
  document.getElementById('runBtn').disabled = !(on && chosen.length>=2);
  if(note) note.textContent = on ? '' : 'Tick "Manuals" above to enable the comparison.';
}
function renderFiles(){
  var ul=document.getElementById('filelist');ul.innerHTML='';
  chosen.forEach(function(f,i){var li=document.createElement('li');
    li.innerHTML='<span class="nm mono">'+f.name+'</span><button class="rm" onclick="removeFile('+i+')" title="Remove">&times;</button>';
    ul.appendChild(li);});
  checkReady();
}

function runTool(){
  if(!manualsOn()){showErr('Please tick "Manuals" to run the comparison.');return;}
  document.getElementById('errBox').innerHTML='';
  var pw=document.getElementById('progressWrap'),pb=document.getElementById('progressBar'),pm=document.getElementById('progressMsg');
  pw.style.display='block';pb.style.width='25%';pm.textContent='Uploading and comparing '+chosen.length+' files\u2026';
  document.getElementById('runBtn').disabled=true;
  var fd=new FormData();chosen.forEach(function(f){fd.append('files',f);});
  fetch('/compare',{method:'POST',body:fd}).then(function(r){return r.json();})
  .then(function(j){pb.style.width='100%';pm.textContent='Done';document.getElementById('runBtn').disabled=false;
    if(!j.ok){showErr(j.error||'Unknown error');pw.style.display='none';return;}
    MODEL=j.model;renderResults();setTimeout(function(){pw.style.display='none';},600);})
  .catch(function(e){showErr('Request failed: '+e);pw.style.display='none';document.getElementById('runBtn').disabled=false;});
}
function showErr(m){document.getElementById('errBox').innerHTML='<div class="err">'+m+'</div>';}
function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/"/g,'&quot;');}
function dispSplit(t){var i=t.lastIndexOf('\u2192');return i>0?[t.slice(0,i).trim(),t.slice(i+1).trim()]:[t,''];}

// majority value in an array; on a tie, the FIRST element (leftmost document) wins
function majority(arr){
  var cnt={},best=null,bn=-1;
  arr.forEach(function(v){cnt[v]=(cnt[v]||0)+1;});
  for(var i=0;i<arr.length;i++){var v=arr[i];if(cnt[v]>bn){bn=cnt[v];best=v;}}
  return best;
}
// which words of 'cw' are part of the longest common subsequence with 'rw'
function lcsMatched(rw,cw){
  var n=rw.length,m=cw.length,dp=[];
  for(var i=0;i<=n;i++){dp.push(new Array(m+1).fill(0));}
  for(i=1;i<=n;i++)for(var j=1;j<=m;j++)
    dp[i][j]=(rw[i-1]===cw[j-1])?dp[i-1][j-1]+1:Math.max(dp[i-1][j],dp[i][j-1]);
  var matched=new Array(m).fill(false); i=n; j=m;
  while(i>0&&j>0){
    if(rw[i-1]===cw[j-1]){matched[j-1]=true;i--;j--;}
    else if(dp[i-1][j]>=dp[i][j-1])i--; else j--;
  }
  return matched;
}
// return HTML for 'cell' with words that differ from reference 'ref' wrapped in red
function diffRed(ref,cell){
  var rw=ref.split(/\s+/).filter(Boolean);
  var cw=cell.split(/\s+/).filter(Boolean);
  var matched=lcsMatched(rw,cw);
  var toks=cell.split(/(\s+)/), wi=0, out='';
  toks.forEach(function(tok){
    if(/^\s+$/.test(tok)||tok===''){out+=tok;return;}
    var ok=matched[wi];wi++;
    out+= ok ? esc(tok) : '<span class="d">'+esc(tok)+'</span>';
  });
  return out;
}

function renderResults(){
  document.getElementById('results').style.display='block';
  document.getElementById('rTitle').textContent=MODEL.task_title||'Comparison results';
  document.getElementById('rTask').textContent=(MODEL.task_no?('TASK '+MODEL.task_no+'  \u00b7  '):'')
    +MODEL.names.length+' documents \u00b7 '+MODEL.rows.length+' criteria \u00b7 '+MODEL.generated;
  var wb=document.getElementById('warnBox');wb.innerHTML='';
  (MODEL.warnings||[]).forEach(function(w){wb.innerHTML+='<div class="warn">'+w+'</div>';});

  // per-doc cards
  var eb=document.getElementById('engines');eb.innerHTML='';
  MODEL.names.forEach(function(nm,gi){var c=MODEL.summary[gi];
    var tot=0;ORDER.forEach(function(k){tot+=c[k];});if(!tot)tot=1;
    function seg(k){return '<span style="width:'+(100*c[k]/tot)+'%;background:'+STCOLOR[k]+'"></span>';}
    var counts=[['MATCH','match'],['PARTIAL','partial'],['THRESHOLD','threshold'],['MISSING','missing'],['UNIQUE','unique']]
      .map(function(p){return '<span><b>'+c[p[0]]+'</b> '+p[1]+'</span>';}).join('');
    var el=document.createElement('div');el.className='ecard';
    el.innerHTML='<h3>'+esc(nm)+'</h3><div class="bar">'+ORDER.map(seg).join('')+'</div><div class="counts">'+counts+'</div>';
    eb.appendChild(el);});

  // pairwise matrix
  var mw=document.getElementById('matrixWrap');
  var h='<table class="matrix"><tr><th></th>'+MODEL.names.map(function(n){return '<th>'+esc(n)+'</th>';}).join('')+'</tr>';
  MODEL.pairs.forEach(function(rowv,i){
    h+='<tr><th>'+esc(MODEL.names[i])+'</th>';
    rowv.forEach(function(v,j){
      if(i===j){h+='<td class="self">\u2014</td>';}
      else{var bg=v===null?'#fff':(v>=90?'var(--c-match)':(v>=70?'var(--c-partial)':'var(--c-miss)'));
        h+='<td style="background:'+bg+'">'+(v===null?'\u2014':v+'%')+'</td>';}
    });h+='</tr>';
  });h+='</table>';mw.innerHTML=h;

  // chips
  var chips=document.getElementById('chips');chips.innerHTML='';
  ORDER.forEach(function(k){var b=document.createElement('span');b.className='chip'+(active[k]?'':' off');
    b.innerHTML='<span class="sw" style="background:'+STCOLOR[k]+'"></span>'+STLABEL[k];
    b.onclick=function(){active[k]=!active[k];b.classList.toggle('off');renderRows();};chips.appendChild(b);});

  // header
  document.getElementById('hrow').innerHTML='<th class="area">Inspection Area</th><th>Criterion</th>'
    +MODEL.names.map(function(n){return '<th>'+esc(n)+'</th>';}).join('');
  document.getElementById('q').oninput=renderRows;
  document.getElementById('dlBtn').onclick=downloadExcel;
  renderSuggestions();
  renderRows();
}

function hlWords(name, shared){
  return name.split(/(\s+)/).map(function(tok){
    var w=tok.toLowerCase().replace(/[^a-z0-9]/g,'');
    if(tok.trim()&&shared.indexOf(w)<0&&!/^(examine|the|and|for|of|a|an|area|part)$/.test(w))
      return '<span class="wd">'+esc(tok)+'</span>';   // the differing bit (e.g. the item number) in red
    return esc(tok);
  }).join('');
}
function renderSuggestions(){
  var sug=MODEL.suggestions||[];
  var det=document.getElementById('matchDetails');
  document.getElementById('matchCount').textContent = sug.length? '('+sug.length+' found)' : '(none)';
  var box=document.getElementById('matchList'); box.innerHTML='';
  if(!sug.length){ box.innerHTML='<div style="font-size:12px;color:var(--muted)">No differently-numbered items detected across the files.</div>';
    document.getElementById('applyMatchBtn').style.display='none'; return; }
  document.getElementById('applyMatchBtn').style.display='';
  sug.forEach(function(s,i){
    var names=s.members.map(function(mm){return '<span class="nm"><span class="fb">'+esc(MODEL.names[mm.file])+'</span>'+hlWords(mm.name,s.shared)+'</span>';}).join('');
    var d=document.createElement('label'); d.className='msug';
    d.innerHTML='<input type="checkbox" data-i="'+i+'"><div class="names">'+names+'</div><span class="sc">'+Math.round(s.score*100)+'%</span>';
    box.appendChild(d);
  });
}
function applyMatches(){
  var picks=[]; 
  document.querySelectorAll('#matchList input[type=checkbox]').forEach(function(cb){
    if(cb.checked){ var s=MODEL.suggestions[+cb.dataset.i]; if(s) picks.push(s.names); }
  });
  var prev=MODEL.aliases||[];
  var aliases=prev.concat(picks);
  var note=document.getElementById('matchNote'); note.textContent='Re-comparing…';
  var btn=document.getElementById('applyMatchBtn'); btn.disabled=true;
  fetch('/recompute',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({token:MODEL.token,aliases:aliases})})
  .then(function(r){return r.json();})
  .then(function(j){ btn.disabled=false;
    if(!j.ok){note.textContent='Error: '+(j.error||'failed');return;}
    var keepAliases=aliases; MODEL=j.model; MODEL.aliases=keepAliases;
    note.textContent='Merged '+aliases.length+' item group(s).';
    renderResults();
  }).catch(function(e){btn.disabled=false;note.textContent='Failed: '+e;});
}

function activeStatuses(){return ORDER.filter(function(k){return active[k];});}

function renderRows(){
  var q=document.getElementById('q').value.toLowerCase().trim(), sset=activeStatuses();
  var tb=document.getElementById('tbody');tb.innerHTML='';var last=null,shown=0;
  MODEL.rows.forEach(function(row){
    var cstat=row.cells.map(function(c){return c.status;});
    if(!cstat.some(function(s){return sset.indexOf(s)>=0;}))return;
    if(q){var hay=(row.area+' '+row.condition+' '+row.criterion+' '+row.cells.map(function(c){return c.text;}).join(' ')).toLowerCase();
      if(hay.indexOf(q)<0)return;}
    var top=row.area!==last;var tr=document.createElement('tr');if(top)tr.className='arowtop';
    var area='<th class="area">'+(top?esc(row.area):'')+(top&&row.merged?'<span class="mergedtag">merged</span>':'')+'</th>';
    var crit='<td class="crit"><span class="mono">'+esc(row.criterion)+'</span>'
      +(row.condition?'<div style="color:#8a97a6;font-size:11px;margin-top:2px">'+esc(row.condition)+'</div>':'')+'</td>';
    // split each cell into main text + disposition
    row.cells.forEach(function(c){var sp=dispSplit(c.text);c._main=sp[0];c._disp=sp[1];});
    // reference = what the majority of present documents say (tie -> leftmost)
    var presentMains=[], presentDisps=[];
    row.cells.forEach(function(c){if(c.status!=='NA'&&c.status!=='MISSING'){presentMains.push(c._main);presentDisps.push(c._disp);}});
    var refMain=presentMains.length?majority(presentMains):null;
    var refDisp=presentDisps.length?majority(presentDisps):null;
    var cells=row.cells.map(function(c){
      var flag=(c.status!=='MATCH'&&c.status!=='NA')?'<span class="flag">'+(STLABEL[c.status]||'')+'</span>':'';
      var mainHtml;
      if(c.status==='NA'||c.status==='MISSING'||refMain===null||c._main===refMain){
        mainHtml='<span class="mono">'+esc(c._main)+'</span>';          // agrees with the set -> plain
      }else{
        mainHtml='<span class="mono">'+diffRed(refMain,c._main)+'</span>'; // differs -> red on the differing words
      }
      var dispHtml='';
      if(c._disp){
        var dred=(c.status!=='NA'&&c.status!=='MISSING'&&refDisp!==null&&c._disp!==refDisp);
        dispHtml='<span class="disp'+(dred?' dred':'')+'">'+esc(c._disp)+'</span>';
      }
      return '<td class="cell st-'+c.status+'" title="'+esc(c.text)+'">'+flag+mainHtml+dispHtml+'</td>';
    }).join('');
    tr.innerHTML=area+crit+cells;tb.appendChild(tr);last=row.area;shown++;
  });
  document.getElementById('dlNote').textContent=shown+' of '+MODEL.rows.length+' rows shown \u2014 this is what downloads';
}

function downloadExcel(){
  var payload={token:MODEL.token,statuses:activeStatuses(),query:document.getElementById('q').value.trim()};
  var btn=document.getElementById('dlBtn');var old=btn.innerHTML;btn.innerHTML='Preparing\u2026';btn.disabled=true;
  fetch('/download',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)})
  .then(function(r){if(!r.ok)return r.json().then(function(j){throw new Error(j.error||'error');});return r.blob();})
  .then(function(blob){var url=URL.createObjectURL(blob);var a=document.createElement('a');a.href=url;
    a.download='CAIRO-Assist_'+(MODEL.task_no||'task')+'.xlsx';document.body.appendChild(a);a.click();
    a.remove();URL.revokeObjectURL(url);btn.innerHTML=old;btn.disabled=false;})
  .catch(function(e){alert('Download failed: '+e.message);btn.innerHTML=old;btn.disabled=false;});
}
</script>
</body></html>"""

# ================================================================= MAIN
LOGS_HTML = r"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>CAIRO-Assist - Usage log</title>
<style>
 body{margin:0;background:#eef2f9;color:#1e293b;font-family:system-ui,"Segoe UI",sans-serif;font-size:13px}
 header{background:linear-gradient(180deg,#f6f8fb,#e9eef6);border-bottom:2px solid #cfd8e6;padding:12px 20px}
 header h1{margin:0;font-size:18px;color:#12203a}
 header p{margin:3px 0 0;color:#5b6675;font-size:12px}
 .bar{padding:12px 20px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
 .bar a{background:#1a4fad;color:#fff;text-decoration:none;border-radius:7px;padding:8px 14px;font-size:13px}
 .bar .path{font-family:Consolas,monospace;font-size:12px;color:#5b6675;background:#fff;border:1px solid #c8d4e8;border-radius:6px;padding:6px 10px}
 .wrap{padding:0 20px 24px}
 .tablebox{background:#fff;border:1px solid #c8d4e8;border-radius:8px;overflow:auto;max-height:78vh}
 table{border-collapse:collapse;width:100%;font-size:12.5px}
 th,td{padding:8px 10px;border-bottom:1px solid #eceef2;text-align:left;vertical-align:top;white-space:nowrap}
 thead th{position:sticky;top:0;background:#1f3fa0;color:#fff;font-weight:600;z-index:2}
 td:nth-child(8){white-space:normal;min-width:280px}
 tbody tr:nth-child(even){background:#f7f9fc}
</style></head><body>
<header><h1>CAIRO-Assist &ndash; Usage log</h1>
<p>Backend audit of who ran the tool, on which machine, and which documents were compared.</p></header>
<div class="bar">
  <a href="/logs.csv">&#8681; Download CSV</a>
  <a href="/" style="background:#fff;color:#1a4fad;border:1px solid #c8d4e8">&larr; Back to tool</a>
  <span class="path">DB: __PATH__</span>
  <span style="color:#5b6675">__COUNT__ most-recent entries</span>
</div>
<div class="wrap"><div class="tablebox">
  <table><thead><tr>__HEAD__</tr></thead><tbody>__ROWS__</tbody></table>
</div></div>
</body></html>"""

def open_browser():
    time.sleep(1.2)
    webbrowser.open('http://127.0.0.1:%d' % PORT)

if __name__ == '__main__':
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='127.0.0.1', port=PORT, debug=False, use_reloader=False)
